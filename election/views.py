import base64
import qrcode.image.svg
import qrcode
from django_otp import verify_token, user_has_device
from django_otp.plugins.otp_totp.models import TOTPDevice
from .forms import CSVUploadForm
from .models import Student
from django.db import connection
from django.shortcuts import redirect, render
import time
import json
import csv
import io

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.db import transaction
from django.db.models import Count
from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.cache import cache
from django.contrib.auth.hashers import make_password

from .models import Student, Election, Position, Candidate, Vote
from .forms import (
    StudentLoginForm, ForcePasswordChangeForm,
    CSVUploadForm, ElectionForm, PositionForm,
    CandidateForm, StudentPasswordChangeForm
)
from .utils import (
    get_active_users,
    create_notification,
    create_bulk_notifications
)


# ═══════════════════════════════════════════════
# AUTH VIEWS
# ═══════════════════════════════════════════════

LOCKOUT_ATTEMPTS = 5
LOCKOUT_WINDOW = 60  # seconds


def login_view(request):

    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        return redirect('student_dashboard')

    form = StudentLoginForm()

    if request.method == 'POST':
        ip = request.META.get('REMOTE_ADDR', 'unknown')
        cache_key = f'login_attempts_{ip}'
        attempts = cache.get(cache_key, 0)

        # ── Rate limit check ──────────────────────────────
        if attempts >= LOCKOUT_ATTEMPTS:
            messages.error(
                request,
                'Too many login attempts. '
                'Please wait 1 minute before trying again.'
            )
            return render(
                request,
                'student/login.html',
                {'form': StudentLoginForm()}
            )

        form = StudentLoginForm(request.POST)
        if form.is_valid():
            admission_number = form.cleaned_data['admission_number']
            password = form.cleaned_data['password']

            user = authenticate(
                request,
                username=str(admission_number),
                password=password
            )

            if user is not None:
                # ── Success — clear lockout ───────────────
                cache.delete(cache_key)
                login(request, user)
                messages.success(
                    request,
                    f'Welcome back, {user.first_name}!'
                )
                return redirect('student_dashboard')

            else:
                # ── Failed — increment counter ────────────
                cache.set(
                    cache_key,
                    attempts + 1,
                    timeout=LOCKOUT_WINDOW
                )
                try:
                    student = Student.objects.get(
                        admission_number=admission_number
                    )
                    if not student.is_active:
                        messages.error(
                            request,
                            'Your account has been deactivated. '
                            'Please contact the administrator.'
                        )
                    else:
                        messages.error(
                            request,
                            'Incorrect password. Please try again.'
                        )
                except Student.DoesNotExist:
                    messages.error(
                        request,
                        'You are not a registered voter. '
                        'Please contact the administrator.'
                    )

    return render(request, 'student/login.html', {'form': form})


def logout_view(request):
    if request.user.is_authenticated:
        # Remove from active users tracker
        cache.delete(f'active_user_{request.user.id}')
        active_ids = cache.get('active_user_ids', set())
        active_ids.discard(request.user.id)
        cache.set('active_user_ids', active_ids, timeout=300)

    logout(request)
    messages.success(
        request, 'You have been logged out successfully.'
    )
    return redirect('login')


def change_password_view(request):

    if not request.user.is_authenticated:
        return redirect('login')

    if request.user.is_staff:
        return redirect('admin_dashboard')

    form = ForcePasswordChangeForm()

    if request.method == 'POST':
        form = ForcePasswordChangeForm(request.POST)
        if form.is_valid():
            request.user.set_password(
                form.cleaned_data['new_password1']
            )
            request.user.password_changed = True
            request.user.save()

            user = authenticate(
                request,
                username=str(request.user.admission_number),
                password=form.cleaned_data['new_password1']
            )
            if user:
                login(request, user)
                create_notification(
                    student=user,
                    title='Password Changed',
                    message=(
                        'Your password has been changed successfully. '
                        'If you did not make this change, contact '
                        'the administrator immediately.'
                    ),
                    notif_type='info',
                )

            messages.success(
                request,
                'Password changed successfully! Welcome to KuraVote.'
            )
            return redirect('student_dashboard')

    return render(
        request,
        'student/change_password.html',
        {'form': form}
    )


# ═══════════════════════════════════════════════
# STUDENT VIEWS
# ═══════════════════════════════════════════════

@login_required(login_url='login')
def student_dashboard(request):

    if request.user.is_staff:
        return redirect('admin_dashboard')

    try:
        election = Election.objects.get(status=Election.STATUS_ACTIVE)
    except Election.DoesNotExist:
        election = None

    voted_positions = []
    if election:
        voted_positions = list(
            Vote.objects.filter(
                student=request.user,
                election=election
            ).values_list('position_id', flat=True)
        )

    if not request.user.is_active:
        messages.warning(
            request,
            'Your account has been deactivated. '
            'You cannot vote. Please contact the administrator.'
        )

    context = {
        'election':        election,
        'voted_positions': voted_positions,
        'total_positions': election.positions.count() if election else 0,
        'votes_cast':      len(voted_positions),
    }
    return render(request, 'student/dashboard.html', context)


@login_required(login_url='login')
def ballot_view(request):

    if request.user.is_staff:
        return redirect('admin_dashboard')

    if not request.user.is_active:
        messages.error(
            request,
            'Your account has been deactivated. '
            'You cannot vote. Please contact the administrator.'
        )
        return redirect('student_dashboard')

    try:
        election = Election.objects.get(status=Election.STATUS_ACTIVE)
    except Election.DoesNotExist:
        messages.error(request, 'No active election at the moment.')
        return redirect('student_dashboard')

    positions = Position.objects.filter(
        election=election
    ).prefetch_related('candidates__student')

    voted_position_ids = list(
        Vote.objects.filter(
            student=request.user,
            election=election
        ).values_list('position_id', flat=True)
    )

    voted_candidate_ids = list(
        Vote.objects.filter(
            student=request.user,
            election=election
        ).values_list('candidate_id', flat=True)
    )

    all_positions_ids = list(
        positions.values_list('id', flat=True)
    )
    all_voted = all(
        pid in voted_position_ids for pid in all_positions_ids
    )

    unvoted_positions = [
        {'id': p.id, 'name': p.position_name}
        for p in positions
        if p.id not in voted_position_ids
    ]

    context = {
        'election':               election,
        'positions':              positions,
        'voted_position_ids':     voted_position_ids,
        'voted_candidate_ids':    voted_candidate_ids,
        'all_voted':              all_voted,
        'total_positions':        len(all_positions_ids),
        'unvoted_positions_json': json.dumps(unvoted_positions),
    }
    return render(request, 'student/ballot.html', context)


@login_required(login_url='login')
def cast_vote(request):

    if request.user.is_staff:
        return redirect('admin_dashboard')

    if not request.user.is_active:
        messages.error(
            request,
            'Your account has been deactivated. '
            'You cannot vote.'
        )
        return redirect('student_dashboard')

    if request.method != 'POST':
        return redirect('ballot')

    try:
        election = Election.objects.get(status=Election.STATUS_ACTIVE)
    except Election.DoesNotExist:
        messages.error(request, 'No active election at the moment.')
        return redirect('student_dashboard')

    selections = {}
    for key, value in request.POST.items():
        if key.startswith('vote_'):
            try:
                position_id = int(key.replace('vote_', ''))
                candidate_id = int(value)
                selections[position_id] = candidate_id
            except (ValueError, TypeError):
                continue

    if not selections:
        messages.error(request, 'No votes were selected.')
        return redirect('ballot')

    errors = []
    saved = 0

    for position_id, candidate_id in selections.items():
        try:
            with transaction.atomic():
                position = Position.objects.get(
                    pk=position_id,
                    election=election
                )
                candidate = Candidate.objects.get(
                    pk=candidate_id,
                    position=position
                )

                already_voted = Vote.objects.select_for_update().filter(
                    student=request.user,
                    position=position,
                    election=election
                ).exists()

                if already_voted:
                    continue

                Vote.objects.create(
                    student=request.user,
                    candidate=candidate,
                    position=position,
                    election=election
                )
                saved += 1

        except (Position.DoesNotExist, Candidate.DoesNotExist):
            errors.append('Invalid selection detected.')
        except Exception:
            errors.append(
                f'Could not save vote for position {position_id}.'
            )

    if saved > 0:
        plural = 's' if saved != 1 else ''
        messages.success(
            request,
            f'Your votes have been submitted successfully! '
            f'{saved} vote{plural} recorded.'
        )
        # Notify student of successful vote
        create_notification(
            student=request.user,
            title='Vote Submitted Successfully',
            message=(
                f'Your {saved} vote{plural} for '
                f'"{election.election_name}" have been '
                f'recorded securely. Thank you for participating!'
            ),
            notif_type='success',
        )

    if errors:
        for error in errors:
            messages.error(request, error)

    return redirect('ballot')


@login_required(login_url='login')
def student_results(request):

    if request.user.is_staff:
        return redirect('admin_dashboard')

    elections = Election.objects.exclude(
        status=Election.STATUS_INACTIVE
    ).order_by('-created_at')

    election_id = request.GET.get('election')
    if election_id:
        try:
            selected_election = elections.get(pk=election_id)
        except Election.DoesNotExist:
            selected_election = elections.first()
    else:
        selected_election = elections.first()

    results = []
    if selected_election:
        positions = Position.objects.filter(
            election=selected_election
        ).prefetch_related('candidates__student')

        for position in positions:
            candidates = sorted(
                position.candidates.all(),
                key=lambda c: c.vote_count,
                reverse=True
            )
            results.append({
                'position':    position,
                'candidates':  candidates,
                'total_votes': position.total_votes,
            })

    context = {
        'elections':         elections,
        'selected_election': selected_election,
        'results':           results,
    }
    return render(request, 'student/results.html', context)


@login_required(login_url='login')
def student_profile(request):

    if request.user.is_staff:
        return redirect('admin_dashboard')

    form = StudentPasswordChangeForm(user=request.user)

    if request.method == 'POST':
        form = StudentPasswordChangeForm(
            user=request.user,
            data=request.POST
        )
        if form.is_valid():
            form.save()
            user = authenticate(
                request,
                username=str(request.user.admission_number),
                password=form.cleaned_data['new_password1']
            )
            if user:
                login(request, user)
            messages.success(
                request,
                'Password changed successfully!'
            )
            return redirect('student_profile')
        else:
            messages.error(
                request,
                'Please fix the errors below.'
            )

    try:
        election = Election.objects.get(status=Election.STATUS_ACTIVE)
        voted_positions = list(
            Vote.objects.filter(
                student=request.user,
                election=election
            ).values_list('position_id', flat=True)
        )
    except Election.DoesNotExist:
        election = None
        voted_positions = []

    context = {
        'form':            form,
        'student':         request.user,
        'election':        election,
        'voted_positions': voted_positions,
    }
    return render(request, 'student/profile.html', context)


@login_required(login_url='login')
def student_notifications(request):

    if request.user.is_staff:
        return redirect('admin_dashboard')

    notifications = request.user.notifications.all()

    request.user.notifications.filter(
        is_read=False
    ).update(is_read=True)

    context = {
        'notifications': notifications,
    }
    return render(
        request,
        'student/notifications.html',
        context
    )


@login_required(login_url='login')
def notifications_api(request):

    if request.user.is_staff:
        return JsonResponse({'count': 0, 'notifications': []})

    unread = request.user.notifications.filter(
        is_read=False
    ).values(
        'id', 'title', 'message', 'notif_type', 'created_at'
    )[:10]

    return JsonResponse({
        'count': request.user.notifications.filter(
            is_read=False
        ).count(),
        'notifications': [
            {
                'id':         n['id'],
                'title':      n['title'],
                'message':    n['message'],
                'type':       n['notif_type'],
                'created_at': n['created_at'].strftime(
                    '%d %b %Y, %H:%M'
                ),
            }
            for n in unread
        ],
    })


# ═══════════════════════════════════════════════
# ADMIN VIEWS
# ═══════════════════════════════════════════════

@staff_member_required(login_url='login')
def admin_dashboard(request):

    total_students = Student.objects.filter(is_staff=False).count()
    total_elections = Election.objects.count()
    total_candidates = Candidate.objects.count()

    try:
        active_election = Election.objects.get(
            status=Election.STATUS_ACTIVE
        )
        total_votes = Vote.objects.filter(
            election=active_election
        ).count()
        voters_count = Vote.objects.filter(
            election=active_election
        ).values('student').distinct().count()
        turnout = round(
            (voters_count / total_students) * 100, 1
        ) if total_students > 0 else 0
    except Election.DoesNotExist:
        active_election = None
        total_votes = 0
        voters_count = 0
        turnout = 0

    recent_elections = Election.objects.order_by('-created_at')[:5]
    active_users = get_active_users()

    context = {
        'total_students':     total_students,
        'total_elections':    total_elections,
        'total_candidates':   total_candidates,
        'active_election':    active_election,
        'total_votes':        total_votes,
        'voters_count':       voters_count,
        'turnout':            turnout,
        'recent_elections':   recent_elections,
        'active_users':       active_users,
        'active_users_count': len(active_users),
    }
    return render(request, 'admin/dashboard.html', context)


# ── Elections ─────────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_elections(request):
    elections = Election.objects.order_by('-created_at')
    return render(request, 'admin/elections.html', {
        'elections': elections
    })


@staff_member_required(login_url='login')
def admin_election_create(request):
    form = ElectionForm()
    if request.method == 'POST':
        form = ElectionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Election created successfully.')
            return redirect('admin_elections')
        else:
            messages.error(request, 'Please fix the errors below.')
    return render(request, 'admin/election_form.html', {
        'form': form, 'action': 'Create'
    })


@staff_member_required(login_url='login')
def admin_election_edit(request, pk):
    election = get_object_or_404(Election, pk=pk)
    form = ElectionForm(instance=election)
    if request.method == 'POST':
        form = ElectionForm(request.POST, instance=election)
        if form.is_valid():
            form.save()
            messages.success(request, 'Election updated successfully.')
            return redirect('admin_elections')
        else:
            messages.error(request, 'Please fix the errors below.')
    return render(request, 'admin/election_form.html', {
        'form': form, 'action': 'Edit', 'election': election
    })


@staff_member_required(login_url='login')
def admin_election_delete(request, pk):
    election = get_object_or_404(Election, pk=pk)
    if request.method == 'POST':
        election.delete()
        messages.success(request, 'Election deleted successfully.')
        return redirect('admin_elections')
    return render(request, 'admin/election_confirm_delete.html', {
        'election': election
    })


@staff_member_required(login_url='login')
def admin_election_activate(request, pk):
    election = get_object_or_404(Election, pk=pk)
    if Election.objects.filter(
        status=Election.STATUS_ACTIVE
    ).exists():
        messages.error(
            request,
            'Another election is already active. '
            'Please close it before activating a new one.'
        )
        return redirect('admin_elections')

    election.status = Election.STATUS_ACTIVE
    election.save()

    # Notify all active students
    students = Student.objects.filter(
        is_staff=False, is_active=True
    )
    create_bulk_notifications(
        students=students,
        title='Election Started',
        message=(
            f'"{election.election_name}" is now open for voting. '
            f'Log in and cast your vote before '
            f'{election.end_date.strftime("%d %B %Y at %I:%M %p")}.'
        ),
        notif_type='info',
    )

    messages.success(
        request,
        f'"{election.election_name}" is now active. '
        f'Voting has started.'
    )
    return redirect('admin_elections')


@staff_member_required(login_url='login')
def admin_election_close(request, pk):
    election = get_object_or_404(Election, pk=pk)
    election.status = Election.STATUS_CLOSED
    election.save()

    # Notify all active students
    students = Student.objects.filter(
        is_staff=False, is_active=True
    )
    create_bulk_notifications(
        students=students,
        title='Election Closed',
        message=(
            f'"{election.election_name}" has ended. '
            f'Results are now available. '
            f'Check the results page to see the outcome.'
        ),
        notif_type='warning',
    )

    messages.success(
        request,
        f'"{election.election_name}" has been closed. '
        f'Voting has ended.'
    )
    return redirect('admin_elections')


# ── Positions ─────────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_positions(request):
    positions = Position.objects.select_related(
        'election'
    ).order_by('election', 'position_name')
    return render(request, 'admin/positions.html', {
        'positions': positions
    })


@staff_member_required(login_url='login')
def admin_position_create(request):
    form = PositionForm()
    if request.method == 'POST':
        form = PositionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Position created successfully.')
            return redirect('admin_positions')
        else:
            messages.error(request, 'Please fix the errors below.')
    return render(request, 'admin/position_form.html', {
        'form': form, 'action': 'Create'
    })


@staff_member_required(login_url='login')
def admin_position_edit(request, pk):
    position = get_object_or_404(Position, pk=pk)
    form = PositionForm(instance=position)
    if request.method == 'POST':
        form = PositionForm(request.POST, instance=position)
        if form.is_valid():
            form.save()
            messages.success(request, 'Position updated successfully.')
            return redirect('admin_positions')
        else:
            messages.error(request, 'Please fix the errors below.')
    return render(request, 'admin/position_form.html', {
        'form': form, 'action': 'Edit', 'position': position
    })


@staff_member_required(login_url='login')
def admin_position_delete(request, pk):
    position = get_object_or_404(Position, pk=pk)
    if request.method == 'POST':
        position.delete()
        messages.success(request, 'Position deleted successfully.')
        return redirect('admin_positions')
    return render(request, 'admin/position_confirm_delete.html', {
        'position': position
    })


# ── Candidates ────────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_candidates(request):
    candidates = Candidate.objects.select_related(
        'student', 'position', 'position__election'
    ).order_by('position', 'student__last_name')
    return render(request, 'admin/candidates.html', {
        'candidates': candidates
    })


@staff_member_required(login_url='login')
def admin_candidate_create(request):
    form = CandidateForm()
    if request.method == 'POST':
        form = CandidateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Candidate registered successfully.'
            )
            return redirect('admin_candidates')
        else:
            messages.error(request, 'Please fix the errors below.')
    return render(request, 'admin/candidate_form.html', {
        'form': form, 'action': 'Register'
    })


@staff_member_required(login_url='login')
def admin_candidate_edit(request, pk):
    candidate = get_object_or_404(Candidate, pk=pk)
    form = CandidateForm(instance=candidate)
    if request.method == 'POST':
        form = CandidateForm(
            request.POST, request.FILES, instance=candidate
        )
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Candidate updated successfully.'
            )
            return redirect('admin_candidates')
        else:
            messages.error(request, 'Please fix the errors below.')
    return render(request, 'admin/candidate_form.html', {
        'form': form, 'action': 'Edit', 'candidate': candidate
    })


@staff_member_required(login_url='login')
def admin_candidate_delete(request, pk):
    candidate = get_object_or_404(Candidate, pk=pk)
    if request.method == 'POST':
        candidate.delete()
        messages.success(request, 'Candidate removed successfully.')
        return redirect('admin_candidates')
    return render(request, 'admin/candidate_confirm_delete.html', {
        'candidate': candidate
    })


# ── Students ──────────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_students(request):
    students = Student.objects.filter(
        is_staff=False
    ).order_by('admission_number')

    query = request.GET.get('q', '')
    if query:
        students = students.filter(
            admission_number__icontains=query
        ) | students.filter(
            first_name__icontains=query
        ) | students.filter(
            last_name__icontains=query
        ) | students.filter(
            email__icontains=query
        )

    # Stats
    all_students = Student.objects.filter(is_staff=False)
    active_count = all_students.filter(is_active=True).count()
    inactive_count = all_students.filter(is_active=False).count()
    default_password_count = all_students.filter(
        password_changed=False
    ).count()

    paginator = Paginator(students, 25)
    page = request.GET.get('page', 1)
    students = paginator.get_page(page)

    return render(request, 'admin/students.html', {
        'students':              students,
        'query':                 query,
        'active_count':          active_count,
        'inactive_count':        inactive_count,
        'default_password_count': default_password_count,
    })


@staff_member_required(login_url='login')
def admin_student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk, is_staff=False)

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()

        if not first_name or not last_name or not email:
            messages.error(request, 'All fields are required.')
        elif Student.objects.filter(
            email=email
        ).exclude(pk=pk).exists():
            messages.error(
                request,
                'That email is already used by another student.'
            )
        else:
            student.first_name = first_name
            student.last_name = last_name
            student.email = email
            student.save()
            messages.success(
                request,
                f'{student.get_full_name()} updated successfully.'
            )
            return redirect('admin_students')

    return render(request, 'admin/student_edit.html', {
        'student': student
    })


@staff_member_required(login_url='login')
def admin_students_bulk_action(request):
    if request.method != 'POST':
        return redirect('admin_students')

    action = request.POST.get('action', '')
    select_all_pages = request.POST.get('select_all_pages', '0') == '1'

    # If select_all_pages, override student_ids with all non-staff students
    if select_all_pages:
        student_ids = list(
            Student.objects.filter(is_staff=False).values_list('id', flat=True)
        )
    else:
        student_ids = request.POST.getlist('student_ids')

    # ── Delete All ────────────────────────────────────────
    if action == 'delete_all':
        from django.contrib.admin.models import LogEntry
        ids = list(
            Student.objects.filter(
                is_staff=False
            ).values_list('id', flat=True)
        )
        LogEntry.objects.filter(user_id__in=ids).delete()
        count = Student.objects.filter(is_staff=False).delete()[0]
        messages.success(
            request,
            f'All {count} students deleted successfully.'
        )
        return redirect('admin_students')

    # ── Validate selection ────────────────────────────────
    if not student_ids:
        messages.error(request, 'No students selected.')
        return redirect('admin_students')

    # Only non-staff students
    students = Student.objects.filter(
        pk__in=student_ids,
        is_staff=False
    )
    count = students.count()

    if count == 0:
        messages.error(request, 'No valid students found.')
        return redirect('admin_students')

    # ── Bulk Activate ─────────────────────────────────────
    if action == 'activate':
        students.update(is_active=True)
        messages.success(
            request,
            f'{count} student(s) activated successfully.'
        )

    # ── Bulk Deactivate ───────────────────────────────────
    elif action == 'deactivate':
        students.update(is_active=False)
        messages.success(
            request,
            f'{count} student(s) deactivated successfully.'
        )

    # ── Bulk Reset Password ───────────────────────────────
    elif action == 'reset_password':
        for student in students:
            student.set_password(str(student.admission_number))
            student.password_changed = False
            student.save(
                update_fields=['password', 'password_changed']
            )
        messages.success(
            request,
            f'Passwords reset for {count} student(s).'
        )

    # ── Bulk Delete ───────────────────────────────────────
    elif action == 'delete':
        from django.contrib.admin.models import LogEntry
        ids = list(students.values_list('id', flat=True))
        LogEntry.objects.filter(user_id__in=ids).delete()
        students.delete()
        messages.success(
            request,
            f'{count} student(s) deleted successfully.'
        )

    else:
        messages.error(request, 'Invalid action.')

    return redirect('admin_students')


@staff_member_required(login_url='login')
def admin_student_create(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        admission_number = request.POST.get('admission_number', '').strip()
        email = request.POST.get('email', '').strip()

        errors = []

        if not first_name:
            errors.append('First name is required.')
        if not last_name:
            errors.append('Last name is required.')
        if not admission_number:
            errors.append('Admission number is required.')
        elif not admission_number.isdigit():
            errors.append('Admission number must be digits only.')
        if not email:
            errors.append('Email is required.')

        if not errors:
            adm = int(admission_number)
            if Student.objects.filter(
                admission_number=adm
            ).exists():
                errors.append(
                    f'Admission number {adm} already exists.'
                )
            elif Student.objects.filter(email=email).exists():
                errors.append(
                    f'Email {email} is already registered.'
                )

        if errors:
            for e in errors:
                messages.error(request, e)
            return redirect('admin_students')

        Student.objects.create(
            admission_number=int(admission_number),
            first_name=first_name,
            last_name=last_name,
            email=email,
            password=make_password(admission_number),
            is_active=True,
            is_staff=False,
            is_superuser=False,
            password_changed=False,
        )
        messages.success(
            request,
            f'{first_name} {last_name} added successfully. '
            f'Default password is their admission number.'
        )
    return redirect('admin_students')


@staff_member_required(login_url='login')
def admin_student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk, is_staff=False)
    if request.method == 'POST':
        name = student.get_full_name()
        student.delete()
        messages.success(request, f'{name} deleted successfully.')
        return redirect('admin_students')
    return render(request, 'admin/student_confirm_delete.html', {
        'student': student
    })


@staff_member_required(login_url='login')
def admin_students_delete_all(request):
    if request.method == 'POST':
        confirm = request.POST.get('confirm_text', '')
        if confirm == 'DELETE ALL':
            Student.objects.filter(is_staff=False).delete()
            messages.success(
                request,
                'All students have been deleted successfully.'
            )
        else:
            messages.error(
                request,
                'Confirmation text did not match. '
                'No students deleted.'
            )
    return redirect('admin_students')


# ── Tuning constants ──────────────────────────────────────────────────────────
# hard row cap — prevents memory bombs     # parallel bcrypt threads; tune to your CPU core count
MAX_CSV_ROWS = 15_000
SQL_BATCH_SIZE = 1_000    # rows per INSERT statement
# ─────────────────────────────────────────────────────────────────────────────


def _bulk_hash_passwords(raw_rows):
    """
    Hash using MD5 — fast and throwaway.
    Students must change password on first login (password_changed=False).
    MD5 is acceptable here because the password is public knowledge
    (it's their admission number) and it expires on first login.
    """
    from django.contrib.auth.hashers import make_password
    return [
        make_password(str(row[0]), hasher='md5')
        for row in raw_rows
    ]


def _bulk_insert_students(rows):
    """
    Low-level parameterised INSERT using Django's raw DB cursor.
    Bypasses the ORM model layer entirely — no Python object
    construction, no signal dispatch, no field validation overhead.

    rows: list of tuples matching the INSERT column order below.
    Returns: number of rows actually inserted.
    """
    if not rows:
        return 0

    # Build a single INSERT … ON DUPLICATE KEY UPDATE (MySQL)
    # or INSERT … ON CONFLICT DO NOTHING (PostgreSQL/SQLite).
    # We detect which DB is in use from the connection vendor.
    vendor = connection.vendor   # 'mysql' | 'postgresql' | 'sqlite'

    columns = (
        'admission_number',
        'first_name',
        'last_name',
        'email',
        'password',
        'password_changed',
        'is_active',
        'is_staff',
        'is_superuser',
        'date_joined',
    )
    col_str = ', '.join(columns)
    placeholder = ', '.join(['%s'] * len(columns))

    inserted = 0

    with connection.cursor() as cursor:
        for batch_start in range(0, len(rows), SQL_BATCH_SIZE):
            batch = rows[batch_start: batch_start + SQL_BATCH_SIZE]

            if vendor == 'mysql':
                # INSERT IGNORE skips duplicate admission_number / email
                # without raising an error — fastest MySQL path.
                sql = (
                    f'INSERT IGNORE INTO students ({col_str}) '
                    f'VALUES ({placeholder})'
                )
            elif vendor == 'postgresql':
                sql = (
                    f'INSERT INTO students ({col_str}) '
                    f'VALUES ({placeholder}) '
                    f'ON CONFLICT (admission_number) DO NOTHING'
                )
            else:
                # SQLite (test runner / dev fallback)
                sql = (
                    f'INSERT OR IGNORE INTO students ({col_str}) '
                    f'VALUES ({placeholder})'
                )

            # executemany sends the whole batch in one round-trip
            cursor.executemany(sql, batch)
            inserted += cursor.rowcount

    return inserted


@staff_member_required(login_url='login')
def admin_upload_csv(request):
    form = CSVUploadForm()

    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():

            t_start = time.perf_counter()

            csv_file = request.FILES['csv_file']
            decoded = csv_file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))

            # ── Step 1: fetch existing keys in ONE query ──────────
            existing_admissions = set(
                Student.objects.values_list('admission_number', flat=True)
            )
            existing_emails = set(
                Student.objects.values_list('email', flat=True)
            )

            # ── Step 2: parse & deduplicate — pure Python, fast ───
            raw_rows = []    # [(admission_number_int, first, last, email)]
            seen_in_csv = set()
            skipped = 0
            errors = []
            row_count = 0

            for row_num, row in enumerate(reader, start=2):
                row_count += 1
                if row_count > MAX_CSV_ROWS:
                    messages.error(
                        request,
                        f'CSV exceeds {MAX_CSV_ROWS} rows. '
                        'Split into smaller files.'
                    )
                    return redirect('admin_students')

                try:
                    admission_number = int(row['admission_number'].strip())
                    first_name = row['first_name'].strip()[:100]
                    last_name = row['last_name'].strip()[:100]
                    email = row['email'].strip()[:254].lower()

                    # Skip if already in DB
                    if admission_number in existing_admissions:
                        skipped += 1
                        continue

                    # Skip duplicate email in DB
                    if email in existing_emails:
                        skipped += 1
                        continue

                    # Skip duplicate within this CSV
                    if admission_number in seen_in_csv:
                        skipped += 1
                        continue

                    seen_in_csv.add(admission_number)
                    raw_rows.append(
                        (admission_number, first_name, last_name, email)
                    )

                except KeyError as e:
                    errors.append(
                        f'Row {row_num}: Missing column {e}. '
                        'Headers must be: admission_number, '
                        'first_name, last_name, email'
                    )
                except (ValueError, TypeError) as e:
                    errors.append(f'Row {row_num}: {e}')

            t_parse = time.perf_counter()

            # ── Step 3: hash passwords in PARALLEL ────────────────
            # bcrypt is CPU-bound but releases the GIL enough that
            # ThreadPoolExecutor gives a real 4–8× speedup on multi-core.
            #
            # We hash str(admission_number) for each student — this is
            # the same default password logic as before, just parallelised.
            # ── Step 3: hash passwords FAST ───────────────────────
            hashed = _bulk_hash_passwords(raw_rows)

            t_hash = time.perf_counter()

            # ── Step 4: assemble final row tuples ─────────────────
            from django.utils import timezone
            now = timezone.now()

            db_rows = [
                (
                    raw_rows[i][0],   # admission_number
                    raw_rows[i][1],   # first_name
                    raw_rows[i][2],   # last_name
                    raw_rows[i][3],   # email
                    hashed[i],        # password  (hashed)
                    False,            # password_changed
                    True,             # is_active
                    False,            # is_staff
                    False,            # is_superuser
                    now,              # date_joined
                )
                for i in range(len(raw_rows))
            ]

            # ── Step 5: raw bulk INSERT ────────────────────────────
            created = _bulk_insert_students(db_rows)

            t_done = time.perf_counter()

            # ── Timing breakdown (shown in success message) ────────
            t_total = t_done - t_start
            t_h_sec = t_hash - t_parse
            t_w_sec = t_done - t_hash

            # ── User feedback ──────────────────────────────────────
            if created > 0:
                messages.success(
                    request,
                    f'Successfully imported {created} student(s) '
                    f'in {t_total:.1f}s '
                    f'(hashing: {t_h_sec:.1f}s, '
                    f'DB write: {t_w_sec:.2f}s).'
                )
            if skipped > 0:
                messages.warning(
                    request,
                    f'{skipped} student(s) skipped — '
                    'already exist or duplicates in CSV.'
                )
            for error in errors[:5]:
                messages.error(request, error)
            if len(errors) > 5:
                messages.error(
                    request,
                    f'...and {len(errors) - 5} more errors. '
                    'Check your CSV format.'
                )

            return redirect('admin_students')

    return render(request, 'admin/upload_csv.html', {'form': form})


@staff_member_required(login_url='login')
def admin_reset_password(request, pk):
    student = get_object_or_404(Student, pk=pk, is_staff=False)
    if request.method == 'POST':
        student.set_password(str(student.admission_number))
        student.password_changed = False
        student.save()
        messages.success(
            request,
            f'Password for {student.get_full_name()} reset to '
            f'their admission number.'
        )
        return redirect('admin_students')
    return render(request, 'admin/student_reset_password.html', {
        'student': student
    })


@staff_member_required(login_url='login')
def admin_toggle_student(request, pk):
    student = get_object_or_404(Student, pk=pk, is_staff=False)
    if request.method == 'POST':
        student.is_active = not student.is_active
        student.save()
        status = 'activated' if student.is_active else 'deactivated'
        messages.success(
            request,
            f'{student.get_full_name()} has been {status}.'
        )
    return redirect('admin_students')


# ── Live Results ──────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_live_results(request, election_id):
    election = get_object_or_404(Election, pk=election_id)
    positions = Position.objects.filter(
        election=election
    ).prefetch_related('candidates__student')

    total_students = Student.objects.filter(is_staff=False).count()
    voters_count = Vote.objects.filter(
        election=election
    ).values('student').distinct().count()
    turnout = round((voters_count / total_students) * 100, 1) \
        if total_students > 0 else 0

    context = {
        'election':       election,
        'positions':      positions,
        'total_students': total_students,
        'voters_count':   voters_count,
        'turnout':        turnout,
    }
    return render(request, 'admin/live_results.html', context)


@staff_member_required(login_url='login')
def admin_results_api(request, election_id):
    cache_key = f'results_api_{election_id}'
    cached = cache.get(cache_key)
    if cached:
        return JsonResponse(cached, safe=False)

    election = get_object_or_404(Election, pk=election_id)
    positions = Position.objects.filter(election=election)

    data = []
    for position in positions:
        candidates = sorted(
            Candidate.objects.filter(
                position=position
            ).select_related('student'),
            key=lambda c: c.vote_count,
            reverse=True
        )
        pos_total = Vote.objects.filter(position=position).count()
        data.append({
            'position_id':   position.id,
            'position_name': position.position_name,
            'total_votes':   pos_total,
            'candidates': [
                {
                    'id':         c.id,
                    'name':       c.student.get_full_name(),
                    'initials':   c.get_initials(),
                    'color':      c.get_avatar_color(),
                    'votes':      c.vote_count,
                    'percentage': c.vote_percentage,
                }
                for c in candidates
            ]
        })

    total_students = Student.objects.filter(is_staff=False).count()
    voters_count = Vote.objects.filter(
        election=election
    ).values('student').distinct().count()
    turnout = round((voters_count / total_students * 100), 1) \
        if total_students > 0 else 0

    result = {
        'election_name':  election.election_name,
        'total_votes':    Vote.objects.filter(election=election).count(),
        'voters_count':   voters_count,
        'total_students': total_students,
        'turnout':        turnout,
        'positions':      data,
    }

    cache.set(cache_key, result, timeout=5)
    return JsonResponse(result)


# ── Reports ───────────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_reports(request, election_id):
    election = get_object_or_404(Election, pk=election_id)
    positions = Position.objects.filter(election=election)

    results = []
    for position in positions:
        candidates = sorted(
            Candidate.objects.filter(
                position=position
            ).select_related('student'),
            key=lambda c: c.vote_count,
            reverse=True
        )
        pos_total = Vote.objects.filter(position=position).count()

        if candidates and pos_total > 0 \
                and candidates[0].vote_count > 0:
            winner = candidates[0]
        else:
            winner = None

        results.append({
            'position':    position,
            'candidates':  candidates,
            'winner':      winner,
            'total_votes': pos_total,
        })

    total_students = Student.objects.filter(is_staff=False).count()
    voters_count = Vote.objects.filter(
        election=election
    ).values('student').distinct().count()
    turnout = round((voters_count / total_students) * 100, 1) \
        if total_students > 0 else 0

    context = {
        'election':       election,
        'results':        results,
        'total_students': total_students,
        'voters_count':   voters_count,
        'turnout':        turnout,
    }
    return render(request, 'admin/reports.html', context)


@staff_member_required(login_url='login')
def export_pdf(request, election_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet

    election = get_object_or_404(Election, pk=election_id)
    positions = Position.objects.filter(election=election)

    if election.status != Election.STATUS_CLOSED:
        messages.error(
            request,
            'Reports can only be exported after '
            'the election is closed.'
        )
        return redirect('admin_reports', election_id=election_id)

    if election.total_votes == 0:
        messages.error(
            request,
            'No votes have been cast in this election.'
        )
        return redirect('admin_reports', election_id=election_id)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(
        Paragraph(
            'KuraVote — Official Election Results',
            styles['Title']
        )
    )
    elements.append(
        Paragraph(election.election_name, styles['Heading2'])
    )
    elements.append(Spacer(1, 6))

    total_students = Student.objects.filter(is_staff=False).count()
    voters_count = Vote.objects.filter(
        election=election
    ).values('student').distinct().count()
    turnout = round((voters_count / total_students) * 100, 1) \
        if total_students > 0 else 0

    elements.append(
        Paragraph(
            f'Total Students: {total_students} | '
            f'Voted: {voters_count} | '
            f'Turnout: {turnout}%',
            styles['Normal']
        )
    )
    elements.append(Spacer(1, 20))

    for position in positions:
        candidates = sorted(
            Candidate.objects.filter(
                position=position
            ).select_related('student'),
            key=lambda c: c.vote_count,
            reverse=True
        )
        pos_total = Vote.objects.filter(position=position).count()

        winner = (
            candidates[0]
            if candidates and pos_total > 0
            and candidates[0].vote_count > 0
            else None
        )

        elements.append(
            Paragraph(position.position_name, styles['Heading3'])
        )
        elements.append(
            Paragraph(
                f'Winner: {winner.student.get_full_name()}'
                if winner else 'No votes cast.',
                styles['Normal']
            )
        )
        elements.append(Spacer(1, 8))

        table_data = [['#', 'Candidate', 'Adm No', 'Votes', '%']]
        for i, c in enumerate(candidates, 1):
            table_data.append([
                str(i),
                c.student.get_full_name(),
                str(c.student.admission_number),
                str(c.vote_count),
                f'{c.vote_percentage}%',
            ])

        table = Table(table_data, colWidths=[30, 180, 100, 60, 60])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0),
             colors.HexColor('#166534')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 10),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            ('BACKGROUND', (0, 1), (-1, 1),
             colors.HexColor('#dcfce7')),
            ('FONTNAME',   (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 2), (-1, -1),
             [colors.white, colors.HexColor('#f0fdf4')]),
            ('GRID',       (0, 0), (-1, -1),
             0.5, colors.HexColor('#d1fae5')),
            ('PADDING',    (0, 0), (-1, -1), 7),
            ('ALIGN',      (2, 1), (-1, -1), 'CENTER'),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 20))

    doc.build(elements)
    buffer.seek(0)

    from django.http import HttpResponse
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; '
        f'filename="kuravote_results_{election_id}.pdf"'
    )
    return response


@staff_member_required(login_url='login')
def export_excel(request, election_id):
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from django.http import HttpResponse

    election = get_object_or_404(Election, pk=election_id)
    positions = Position.objects.filter(election=election)

    if election.status != Election.STATUS_CLOSED:
        messages.error(
            request,
            'Reports can only be exported after '
            'the election is closed.'
        )
        return redirect('admin_reports', election_id=election_id)

    if election.total_votes == 0:
        messages.error(
            request,
            'No votes have been cast in this election.'
        )
        return redirect('admin_reports', election_id=election_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Election Results'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(
        start_color='166534', end_color='166534',
        fill_type='solid'
    )
    winner_fill = PatternFill(
        start_color='dcfce7', end_color='dcfce7',
        fill_type='solid'
    )
    winner_font = Font(bold=True, color='166534', size=10)
    alt_fill = PatternFill(
        start_color='f0fdf4', end_color='f0fdf4',
        fill_type='solid'
    )
    center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='d1fae5'),
        right=Side(style='thin', color='d1fae5'),
        top=Side(style='thin', color='d1fae5'),
        bottom=Side(style='thin', color='d1fae5'),
    )

    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f'KuraVote — {election.election_name}'
    title_cell.font = Font(bold=True, size=14, color='166534')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    total_students = Student.objects.filter(is_staff=False).count()
    voters_count = Vote.objects.filter(
        election=election
    ).values('student').distinct().count()
    turnout = round((voters_count / total_students) * 100, 1) \
        if total_students > 0 else 0

    ws.merge_cells('A2:F2')
    summary_cell = ws['A2']
    summary_cell.value = (
        f'Total Students: {total_students}  |  '
        f'Voted: {voters_count}  |  '
        f'Turnout: {turnout}%'
    )
    summary_cell.font = Font(italic=True, size=10,
                             color='64748b')
    summary_cell.alignment = center
    ws.row_dimensions[2].height = 20

    current_row = 4

    for position in positions:
        candidates = sorted(
            Candidate.objects.filter(
                position=position
            ).select_related('student'),
            key=lambda c: c.vote_count,
            reverse=True
        )
        pos_total = Vote.objects.filter(position=position).count()

        winner = (
            candidates[0]
            if candidates and pos_total > 0
            and candidates[0].vote_count > 0
            else None
        )

        ws.merge_cells(f'A{current_row}:F{current_row}')
        pos_cell = ws.cell(
            row=current_row, column=1,
            value=position.position_name.upper()
        )
        pos_cell.font = Font(bold=True, color='FFFFFF', size=11)
        pos_cell.fill = PatternFill(
            start_color='14532d', end_color='14532d',
            fill_type='solid'
        )
        pos_cell.alignment = center
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        if winner:
            ws.merge_cells(f'A{current_row}:F{current_row}')
            win_cell = ws.cell(
                row=current_row, column=1,
                value=f'Winner: {winner.student.get_full_name()}'
            )
            win_cell.font = Font(
                bold=True, color='166534',
                size=10, italic=True
            )
            win_cell.fill = winner_fill
            win_cell.alignment = Alignment(
                horizontal='left', vertical='center', indent=1
            )
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        headers = [
            '#', 'Candidate Name',
            'Admission No', 'Votes', '%', 'Status'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(
                row=current_row, column=col, value=header
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for i, c in enumerate(candidates, 1):
            is_winner = winner and c.id == winner.id
            row_data = [
                i,
                c.student.get_full_name(),
                c.student.admission_number,
                c.vote_count,
                f'{c.vote_percentage}%',
                'WINNER' if is_winner else '',
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(
                    row=current_row, column=col, value=value
                )
                cell.border = thin_border
                cell.alignment = (
                    center if col != 2 else
                    Alignment(
                        horizontal='left',
                        vertical='center',
                        indent=1
                    )
                )
                if is_winner:
                    cell.font = winner_font
                    cell.fill = winner_fill
                elif i % 2 == 0:
                    cell.fill = alt_fill
            ws.row_dimensions[current_row].height = 16
            current_row += 1

        current_row += 1

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = (
        f'attachment; '
        f'filename="kuravote_results_{election_id}.xlsx"'
    )
    return response


# ── Audit Log ─────────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_audit_log(request):
    from django.contrib.admin.models import LogEntry
    logs = LogEntry.objects.select_related(
        'user', 'content_type'
    ).order_by('-action_time')
    paginator = Paginator(logs, 50)
    page = request.GET.get('page', 1)
    logs = paginator.get_page(page)
    return render(request, 'admin/audit_log.html', {'logs': logs})


@staff_member_required(login_url='login')
def admin_active_users_api(request):
    users = get_active_users()
    return JsonResponse({
        'count': len(users),
        'users': users,
    })


# ── Settings ──────────────────────────────────────────────

@staff_member_required(login_url='login')
def admin_settings(request):
    return render(request, 'admin/settings.html')


# ═══════════════════════════════════════════════
# OTP VIEWS
# ═══════════════════════════════════════════════


@staff_member_required(login_url='login')
def admin_setup_otp(request):
    """First-time OTP device registration for admin."""
    user = request.user

    # Already has a device — go to verify
    if user_has_device(user):
        return redirect('admin_verify_otp')

    # Create an unconfirmed device
    device, created = TOTPDevice.objects.get_or_create(
        user=user,
        name='KuraVote Admin',
        defaults={'confirmed': False}
    )

    if request.method == 'POST':
        token = request.POST.get('token', '').strip()
        if device.verify_token(token):
            device.confirmed = True
            device.save()
            messages.success(
                request,
                'Two-factor authentication enabled successfully!'
            )
            return redirect('admin_dashboard')
        else:
            messages.error(request, 'Invalid code. Please try again.')

    # Generate QR code for Google Authenticator / Authy
    otp_url = device.config_url
    qr = qrcode.make(otp_url)
    buf = io.BytesIO()
    qr.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    return render(request, 'admin/setup_otp.html', {
        'device':   device,
        'qr_b64':   qr_b64,
        'otp_url':  otp_url,
    })


@staff_member_required(login_url='login')
def admin_verify_otp(request):
    """Per-session OTP verification."""
    if request.user.is_verified():
        return redirect('admin_dashboard')

    if request.method == 'POST':
        token = request.POST.get('token', '').strip()

        # Find the user's confirmed device
        device = TOTPDevice.objects.filter(
            user=request.user,
            confirmed=True
        ).first()

        if device and device.verify_token(token):
            # Mark session as OTP-verified
            from django_otp import login as otp_login
            otp_login(request, device)
            messages.success(request, 'Verified. Welcome.')
            return redirect(
                request.GET.get('next', 'admin_dashboard')
            )
        else:
            messages.error(
                request,
                'Invalid or expired code. Codes refresh every 30 seconds.'
            )

    return render(request, 'admin/verify_otp.html')
